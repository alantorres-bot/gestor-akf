"""Controle do passivo "por fora" (diferença entre borderô por dentro e por fora).

Importa o histórico da planilha de controle (aba Planilha2) e mantém o saldo.

Colunas da planilha:
  DATA | BASE 01 (nº borderô por dentro) | BASE 02 (nº borderô por fora) |
  VALOR (lançamento: diferença + juros/multa avulsos) | PAGAMENTOS (abatimentos) |
  SALDO | OBSERVAÇÃO

IMPORTANTE (substância econômica): o VALOR é tratado como CUSTO FINANCEIRO.
Estornos/devoluções (ex.: "DEVOLUÇÃO", "ESTORNO") NÃO são pagamento de fato —
ficam em coluna própria para não distorcer o total quitado, conforme o problema
apontado na planilha original (linha de devolução somando dentro de PAGAMENTOS).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from typing import Optional

from .numeros import parse_valor_br
from .util import normalizar_texto, parse_data_br

Z = Decimal("0.00")

# Palavras na OBSERVAÇÃO que indicam estorno/devolução (não é quitação real).
TERMOS_ESTORNO = ("devolu", "estorno", "estornado")


@dataclass
class LancamentoPorFora:
    data: Optional[date]
    base01: str          # nº do borderô por dentro (8xxx)
    base02: str          # nº do borderô por fora (7xxx)
    valor: Decimal       # lançamento de custo (positivo aumenta o passivo)
    pagamento: Decimal   # abatimento real (reduz o passivo)
    estorno: Decimal     # devolução/estorno (separado dos pagamentos)
    observacao: str
    origem_linha: int

    @property
    def eh_estorno(self) -> bool:
        return self.estorno != Z


@dataclass
class ResumoPorFora:
    lancamentos: list[LancamentoPorFora] = field(default_factory=list)
    total_gerado: Decimal = Z       # soma dos VALORES (custo financeiro gerado)
    total_pago: Decimal = Z         # soma dos PAGAMENTOS reais
    total_estornado: Decimal = Z    # soma das devoluções/estornos
    saldo: Decimal = Z              # gerado - pago + estornado

    @property
    def qtd(self) -> int:
        return len(self.lancamentos)


def _eh_estorno(observacao: str) -> bool:
    o = normalizar_texto(observacao)
    return any(termo in o for termo in TERMOS_ESTORNO)


def importar_planilha_por_fora(
    caminho: str,
    aba: str = "Planilha2",
) -> ResumoPorFora:
    """Lê a planilha Excel de controle e devolve o histórico + saldo.

    Detecta a linha de cabeçalho (que contém DATA/VALOR/SALDO) automaticamente,
    então tolera a presença do título "CONTROLE DE JUROS" acima.
    """
    import openpyxl

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        # cai na primeira aba que tenha dados
        aba = wb.sheetnames[-1]
    ws = wb[aba]
    linhas = list(ws.iter_rows(values_only=True))

    # acha a linha de cabeçalho
    hdr_idx = None
    col = {}
    for i, row in enumerate(linhas):
        norm = [normalizar_texto(c) if c is not None else "" for c in row]
        if "data" in norm and any("valor" in n for n in norm):
            hdr_idx = i
            for j, n in enumerate(norm):
                if n == "data":
                    col["data"] = j
                elif n.startswith("base 01") or n == "base 01" or n.replace(" ", "") == "base01":
                    col["base01"] = j
                elif n.replace(" ", "") == "base02":
                    col["base02"] = j
                elif n.startswith("valor"):
                    col.setdefault("valor", j)
                elif n.startswith("pagamento"):
                    col["pagamento"] = j
                elif n.startswith("saldo"):
                    col["saldo"] = j
                elif n.startswith("observa"):
                    col["observacao"] = j
            break

    if hdr_idx is None:
        raise ValueError("Não encontrei o cabeçalho (DATA/VALOR) na planilha.")

    def cell(row, campo):
        j = col.get(campo)
        return row[j] if j is not None and j < len(row) else None

    resumo = ResumoPorFora()
    for n, row in enumerate(linhas[hdr_idx + 1:], start=hdr_idx + 2):
        if row is None or not any(c is not None and str(c).strip() for c in row):
            continue
        data = parse_data_br(cell(row, "data"))
        valor = parse_valor_br(cell(row, "valor")) or Z
        pgto_bruto = parse_valor_br(cell(row, "pagamento")) or Z
        obs = str(cell(row, "observacao") or "").strip()
        base01 = str(cell(row, "base01") or "").strip()
        base02 = str(cell(row, "base02") or "").strip()

        # pula linhas totalmente vazias (sem data, sem valores)
        if data is None and valor == Z and pgto_bruto == Z and not base01 and not base02:
            continue

        # separa estorno de pagamento real
        if pgto_bruto != Z and _eh_estorno(obs):
            estorno, pagamento = pgto_bruto, Z
        else:
            estorno, pagamento = Z, pgto_bruto

        resumo.lancamentos.append(LancamentoPorFora(
            data=data, base01=base01, base02=base02, valor=valor,
            pagamento=pagamento, estorno=estorno, observacao=obs, origem_linha=n,
        ))
        resumo.total_gerado += valor
        resumo.total_pago += pagamento
        resumo.total_estornado += estorno

    resumo.saldo = (resumo.total_gerado - resumo.total_pago + resumo.total_estornado).quantize(
        Decimal("0.01")
    )
    return resumo
